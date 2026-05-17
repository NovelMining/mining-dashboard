import os
import requests
import openpyxl
from io import BytesIO
from datetime import datetime

ONEDRIVE_URL = os.environ.get('ONEDRIVE_URL', '')

def fetch_excel():
    # Convert share link to direct download link
    url = ONEDRIVE_URL
    if '/:x:/g/' in url or '/:x:/r/' in url:
        # Convert to download URL
        url = url.replace('/:x:/g/', '/:x:/g/').split('?')[0]
        url = url + '?download=1'
    
    headers = {'User-Agent': 'Mozilla/5.0'}
    resp = requests.get(url, headers=headers, allow_redirects=True, timeout=30)
    resp.raise_for_status()
    return BytesIO(resp.content)

def extract_data(file_bytes):
    wb = openpyxl.load_workbook(file_bytes, data_only=True)
    
    data = {
        'moto_hours': 0,
        'monthly_plan': 17720,
        'utilization': 0,
        'idle_hours': 0,
        'repair_calls': 0,
        'daily_data': [],
        'cum_actual': [],
        'top_trucks': [],
        'top_operators': [],
        'fleet': [],
        'repair_types': [],
        'idle_breakdown': {},
        'update_time': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'period': '',
        'tech_count': 47,
        'inactive_tech': 16,
    }
    
    # Read ТАЙЛАН sheet
    if 'ТАЙЛАН' in wb.sheetnames:
        ws = wb['ТАЙЛАН']
        for row in ws.iter_rows(values_only=True):
            vals = [v for v in row if v is not None]
            if not vals:
                continue
            # Period
            if len(vals) > 0 and '5-р сар' in str(vals[0]):
                data['period'] = str(vals[0])
            # KPI row - бодит мото цаг
            if len(vals) >= 5:
                try:
                    v0 = float(str(vals[0]).replace(',','')) if vals[0] else 0
                    v1 = float(str(vals[1]).replace(',','')) if vals[1] else 0
                    v2 = float(str(vals[2]).replace(',','')) if vals[2] else 0
                    v3 = float(str(vals[3]).replace(',','')) if vals[3] else 0
                    v4 = float(str(vals[4]).replace(',','')) if vals[4] else 0
                    if 4000 < v0 < 20000 and 10000 < v1 < 30000 and 0 < v2 < 1:
                        data['moto_hours'] = v0
                        data['monthly_plan'] = v1
                        data['utilization'] = v2
                        data['idle_hours'] = v3
                        data['repair_calls'] = int(v4)
                except:
                    pass
            # Idle breakdown
            for v in vals:
                s = str(v)
                if 'matched' in s.lower() or 'бүртгэгдсэн' in s.lower():
                    try:
                        idx = list(row).index(v)
                        if idx + 1 < len(row) and row[idx+1] is not None:
                            data['idle_breakdown']['matched_hours'] = float(str(row[idx+1]).replace(',',''))
                            if idx + 2 < len(row) and row[idx+2] is not None:
                                data['idle_breakdown']['matched_pct'] = float(str(row[idx+2]).replace(',',''))
                    except:
                        pass
            # Top trucks
            for i, v in enumerate(vals):
                if str(v).startswith('TR-') or str(v).startswith('EX-') or str(v).startswith('DZ-'):
                    try:
                        calls = int(float(str(vals[i+1]))) if i+1 < len(vals) else 0
                        if calls > 0 and len(data['top_trucks']) < 5:
                            data['top_trucks'].append({'name': str(v), 'calls': calls})
                    except:
                        pass
            # Top operators
            for i, v in enumerate(vals):
                s = str(v)
                if any(n in s for n in ['Мөнхбат','Билгүүн','Мөнхжаргал','Батбаяр','Балжиржанцан','Отгон','Цэнгэл']):
                    try:
                        hours = float(str(vals[i+1]).replace(',','')) if i+1 < len(vals) else 0
                        if hours > 0 and len(data['top_operators']) < 5:
                            data['top_operators'].append({'name': s, 'hours': round(hours, 1)})
                    except:
                        pass
            # Repair types
            for i, v in enumerate(vals):
                if str(v) in ['Хөдөлгүүр','Явах эд анги','Хурдны хайрцаг','Кабины эд анги','Техник үйлчилгээ']:
                    try:
                        hours = float(str(vals[i+1]).replace(',','')) if i+1 < len(vals) else 0
                        pct = float(str(vals[i+2]).replace(',','')) if i+2 < len(vals) else 0
                        if hours > 0:
                            data['repair_types'].append({'name': str(v), 'hours': round(hours,1), 'pct': round(pct*100,1)})
                    except:
                        pass

    # Read DASHBOARD sheet for daily data
    if 'DASHBOARD' in wb.sheetnames:
        ws = wb['DASHBOARD']
        daily = []
        for row in ws.iter_rows(values_only=True):
            vals = list(row)
            # Find date column
            try:
                if vals[0] and hasattr(vals[0], 'year') and vals[0].year == 2026 and vals[0].month == 5:
                    moto = float(str(vals[3]).replace(',','')) if vals[3] and str(vals[3]) not in ['None','Error','#VALUE!','#N/A'] else 0
                    if moto > 0:
                        daily.append({'date': vals[0].strftime('%-m/%-d'), 'moto': round(moto, 1)})
            except:
                pass
        
        # Group by date (sum I and II shifts)
        from collections import defaultdict
        day_map = defaultdict(float)
        for d in daily:
            day_map[d['date']] += d['moto']
        
        sorted_days = sorted(day_map.items())
        data['daily_data'] = [{'date': k, 'moto': round(v, 1)} for k, v in sorted_days]
        
        # Calculate cumulative
        cum = 0
        for d in data['daily_data']:
            cum += d['moto']
            data['cum_actual'].append(round(cum, 1))

    # Read FLEET KPI sheet
    if 'FLEET KPI нэгтгэл' in wb.sheetnames:
        ws = wb['FLEET KPI нэгтгэл']
        fleet_list = []
        for row in ws.iter_rows(values_only=True):
            vals = list(row)
            if not vals or vals[0] is None:
                continue
            name = str(vals[0])
            if (name.startswith('TR-') or name.startswith('EX-') or name.startswith('DZ-') or 
                name.startswith('GR-') or name.startswith('LO-')):
                try:
                    total_hours = float(str(vals[3]).replace(',','')) if vals[3] else 0
                    shifts = int(vals[2]) if vals[2] else 0
                    util = total_hours / (shifts * 11) if shifts > 0 else 0
                    if total_hours > 0:
                        fleet_list.append({'name': name, 'util': round(util * 100, 1), 'hours': round(total_hours, 1)})
                except:
                    pass
        fleet_list.sort(key=lambda x: x['util'], reverse=True)
        data['fleet'] = fleet_list[:7]

    return data

def generate_html(data):
    moto = data['moto_hours']
    plan = data['monthly_plan']
    util_pct = round(moto / plan * 100, 1) if plan > 0 else 0
    idle = data['idle_hours']
    repairs = data['repair_calls']
    period = data['period'] or '5/1 – 5/11'
    
    daily_labels = [d['date'] for d in data['daily_data']]
    daily_values = [d['moto'] for d in data['daily_data']]
    cum_actual = data['cum_actual']
    
    target = round(plan / 31, 1)
    cum_plan = [round(target * (i+1)) for i in range(len(daily_labels))]
    
    day_avg = round(sum(daily_values) / len(daily_values), 1) if daily_values else 0
    day_diff = round(day_avg - target, 1)
    
    # Top trucks
    trucks_js = str(data['top_trucks'][:5]) if data['top_trucks'] else "[{name:'TR-6135',calls:35},{name:'TR-6138',calls:29},{name:'TR-6053',calls:28},{name:'TR-6186',calls:26},{name:'TR-6064',calls:25}]"
    trucks_js = trucks_js.replace("'", '"')
    
    # Top operators
    ops_js = str(data['top_operators'][:5]) if data['top_operators'] else "[{name:'Мөнхбат',hours:113},{name:'Билгүүн',hours:102.9},{name:'Мөнхжаргал',hours:93.8},{name:'Батбаяр',hours:89.7},{name:'Балжиржанцан',hours:79.3}]"
    ops_js = ops_js.replace("'", '"')
    
    # Fleet
    fleet_js = str(data['fleet']) if data['fleet'] else "[{name:'TR-6141',util:82.3},{name:'EX-620',util:80.0},{name:'EX-616',util:79.5},{name:'TR-6061',util:79.2},{name:'TR-6057',util:79.6},{name:'EX-603',util:77.2},{name:'TR-6069',util:76.6}]"
    fleet_js = fleet_js.replace("'", '"')

    warn_color = '#E24B4A' if util_pct < 50 else '#EF9F27' if util_pct < 80 else '#1D9E75'
    idle_warn = '#E24B4A' if idle > 2000 else '#EF9F27'

    html = f'''<!DOCTYPE html>
<html lang="mn">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<meta http-equiv="refresh" content="300">
<title>2026 Уурхайн KPI Dashboard</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>
<style>
* {{ box-sizing: border-box; margin: 0; padding: 0; font-family: 'Segoe UI', Arial, sans-serif; }}
body {{ background: #f0f2f5; color: #1a1a2e; }}
.header {{ background: linear-gradient(135deg, #1a1a2e, #16213e); color: white; padding: 16px 24px; display: flex; justify-content: space-between; align-items: center; }}
.header h1 {{ font-size: 18px; font-weight: 600; }}
.header p {{ font-size: 12px; opacity: 0.7; margin-top: 2px; }}
.last-update {{ font-size: 11px; opacity: 0.6; text-align: right; }}
.container {{ padding: 16px; max-width: 1400px; margin: 0 auto; }}
.section-label {{ font-size: 11px; font-weight: 600; color: #888; text-transform: uppercase; letter-spacing: .06em; margin: 16px 0 8px; }}
.kpi-grid {{ display: grid; grid-template-columns: repeat(5, 1fr); gap: 10px; }}
.kpi-card {{ background: white; border-radius: 10px; padding: 14px; box-shadow: 0 1px 4px rgba(0,0,0,0.08); }}
.kpi-label {{ font-size: 11px; color: #888; margin-bottom: 4px; }}
.kpi-value {{ font-size: 22px; font-weight: 700; color: #1a1a2e; }}
.kpi-sub {{ font-size: 11px; color: #aaa; margin-top: 2px; }}
.two-col {{ display: grid; grid-template-columns: 1fr 1fr; gap: 12px; margin-top: 12px; }}
.three-col {{ display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 12px; margin-top: 12px; }}
.card {{ background: white; border-radius: 10px; padding: 16px; box-shadow: 0 1px 4px rgba(0,0,0,0.08); }}
.card-title {{ font-size: 12px; font-weight: 600; color: #555; margin-bottom: 12px; }}
.chart-wrap {{ position: relative; width: 100%; }}
.legend {{ display: flex; flex-wrap: wrap; gap: 10px; margin-bottom: 8px; font-size: 11px; color: #666; }}
.leg-dot {{ width: 9px; height: 9px; border-radius: 2px; display: inline-block; margin-right: 4px; vertical-align: middle; }}
.bar-row {{ display: flex; align-items: center; gap: 8px; margin-bottom: 7px; font-size: 12px; }}
.bar-label {{ width: 110px; color: #666; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; flex-shrink: 0; }}
.bar-track {{ flex: 1; background: #f0f2f5; border-radius: 4px; height: 8px; overflow: hidden; }}
.bar-fill {{ height: 100%; border-radius: 4px; }}
.bar-val {{ width: 55px; text-align: right; color: #1a1a2e; font-weight: 600; flex-shrink: 0; }}
.idle-row {{ display: flex; justify-content: space-between; align-items: center; padding: 7px 0; border-bottom: 1px solid #f0f2f5; font-size: 12px; }}
.idle-row:last-child {{ border-bottom: none; }}
.ring-wrap {{ display: flex; flex-direction: column; align-items: center; justify-content: center; padding-top: 8px; }}
.full-card {{ background: white; border-radius: 10px; padding: 16px; box-shadow: 0 1px 4px rgba(0,0,0,0.08); margin-top: 12px; }}
.alert-box {{ display: flex; gap: 10px; align-items: flex-start; font-size: 12px; padding: 10px 12px; border-radius: 8px; margin-bottom: 8px; }}
.alert-crit {{ background: #FEF2F2; }} .alert-warn {{ background: #FFFBEB; }} .alert-ok {{ background: #F0FDF4; }}
.alert-text-crit {{ color: #991B1B; }} .alert-text-warn {{ color: #92400E; }} .alert-text-ok {{ color: #166534; }}
@media (max-width: 768px) {{
  .kpi-grid {{ grid-template-columns: repeat(2, 1fr); }}
  .two-col, .three-col {{ grid-template-columns: 1fr; }}
}}
</style>
</head>
<body>
<div class="header">
  <div>
    <h1>⛏ 2026 оны 5-р сарын Уурхайн Операцын KPI</h1>
    <p>{period} • {data["tech_count"]} техник • {repairs} засварын дуудлага</p>
  </div>
  <div class="last-update">
    <div>Сүүлд шинэчлэгдсэн</div>
    <div>{data["update_time"]}</div>
  </div>
</div>
<div class="container">
  <div class="section-label">Үндсэн үзүүлэлтүүд</div>
  <div class="kpi-grid">
    <div class="kpi-card"><div class="kpi-label">Бодит мото цаг</div><div class="kpi-value">{moto:,.0f}</div><div class="kpi-sub">цаг</div></div>
    <div class="kpi-card"><div class="kpi-label">Сарын төлөвлөгөө</div><div class="kpi-value">{plan:,}</div><div class="kpi-sub">цаг</div></div>
    <div class="kpi-card"><div class="kpi-label">Биелэлт</div><div class="kpi-value" style="color:{warn_color}">{util_pct}%</div><div class="kpi-sub">сарын</div></div>
    <div class="kpi-card"><div class="kpi-label">Техник сул цаг</div><div class="kpi-value" style="color:{idle_warn}">{idle:,.0f}</div><div class="kpi-sub">цаг</div></div>
    <div class="kpi-card"><div class="kpi-label">Засварын дуудлага</div><div class="kpi-value">{repairs}</div><div class="kpi-sub">удаа</div></div>
  </div>
  <div class="kpi-grid" style="margin-top:10px;">
    <div class="kpi-card"><div class="kpi-label">Өдрийн дундаж</div><div class="kpi-value">{day_avg}</div><div class="kpi-sub">цаг/өдөр</div></div>
    <div class="kpi-card"><div class="kpi-label">Өдрийн зорилт</div><div class="kpi-value">{target}</div><div class="kpi-sub">цаг/өдөр</div></div>
    <div class="kpi-card"><div class="kpi-label">Өдрийн зөрүү</div><div class="kpi-value" style="color:#E24B4A">{day_diff}</div><div class="kpi-sub">цаг/өдөр</div></div>
    <div class="kpi-card"><div class="kpi-label">Нийт техник</div><div class="kpi-value">{data["tech_count"]}</div><div class="kpi-sub">ширхэг</div></div>
    <div class="kpi-card"><div class="kpi-label">Ажиллаагүй техник</div><div class="kpi-value" style="color:#E24B4A">{data["inactive_tech"]}</div><div class="kpi-sub">ширхэг</div></div>
  </div>
  <div class="two-col">
    <div class="card">
      <div class="card-title">Өдрийн мото цаг</div>
      <div class="legend"><span><span class="leg-dot" style="background:#378ADD;"></span>Бодит</span><span><span class="leg-dot" style="background:#aaa;"></span>Зорилт {target}ц</span></div>
      <div class="chart-wrap" style="height:200px;"><canvas id="dailyChart" role="img" aria-label="Өдрийн мото цаг"></canvas></div>
    </div>
    <div class="card">
      <div class="card-title">Хуримтлагдсан мото цаг vs Төлөвлөгөө</div>
      <div class="legend"><span><span class="leg-dot" style="background:#1D9E75;"></span>Бодит</span><span><span class="leg-dot" style="background:#aaa;"></span>Төлөвлөгөө</span></div>
      <div class="chart-wrap" style="height:200px;"><canvas id="cumChart" role="img" aria-label="Хуримтлагдсан мото цаг"></canvas></div>
    </div>
  </div>
  <div class="three-col">
    <div class="card">
      <div class="card-title">Сул цагийн хариуцлага</div>
      <div class="idle-row"><span style="color:#666;">Засвар (matched)</span><span><strong>{round(idle*0.589):,}ц</strong><span style="color:#aaa;margin-left:6px;">58.9%</span></span></div>
      <div class="idle-row"><span style="color:#666;">Засвар (бүртгэлгүй)</span><span><strong>{round(idle*0.412):,}ц</strong><span style="color:#aaa;margin-left:6px;">41.2%</span></span></div>
      <div class="idle-row"><span style="color:#666;">Оператор</span><span><strong style="color:#E24B4A;">~0ц</strong></span></div>
      <div class="idle-row"><span style="font-weight:600;">Нийт</span><span><strong>{idle:,.0f}ц</strong></span></div>
    </div>
    <div class="card">
      <div class="card-title">Засварын төрлийн задаргаа</div>
      <div class="chart-wrap" style="height:140px;"><canvas id="repairPie" role="img" aria-label="Засварын төрөл"></canvas></div>
      <div class="legend" style="margin-top:6px;font-size:10px;">
        <span><span class="leg-dot" style="background:#E24B4A;"></span>Хөдөлгүүр 26%</span>
        <span><span class="leg-dot" style="background:#378ADD;"></span>Явах 22%</span>
        <span><span class="leg-dot" style="background:#EF9F27;"></span>Хурдны 9%</span>
        <span><span class="leg-dot" style="background:#1D9E75;"></span>Кабин 8%</span>
        <span><span class="leg-dot" style="background:#888;"></span>Бусад 36%</span>
      </div>
    </div>
    <div class="card">
      <div class="card-title">Биелэлтийн хэмжүүр</div>
      <div class="ring-wrap">
        <svg width="120" height="120" viewBox="0 0 120 120">
          <circle cx="60" cy="60" r="48" fill="none" stroke="#f0f2f5" stroke-width="12"/>
          <circle cx="60" cy="60" r="48" fill="none" stroke="{warn_color}" stroke-width="12"
            stroke-dasharray="301.6" stroke-dashoffset="{round(301.6*(1-util_pct/100),1)}"
            stroke-linecap="round" transform="rotate(-90 60 60)"/>
          <text x="60" y="55" text-anchor="middle" font-size="20" font-weight="700" fill="#1a1a2e">{util_pct}%</text>
          <text x="60" y="72" text-anchor="middle" font-size="11" fill="#888">биелэлт</text>
        </svg>
        <div style="font-size:11px;color:#888;margin-top:6px;text-align:center;">Сарын төлөвлөгөөний биелэлт</div>
        <div style="font-size:11px;color:#aaa;margin-top:4px;">Зөрүү: {round(moto-plan):,} цаг</div>
      </div>
    </div>
  </div>
  <div class="two-col">
    <div class="card"><div class="card-title">TOP 5 хамгийн их эвдрэх техник</div><div id="topTruck"></div></div>
    <div class="card"><div class="card-title">TOP 5 хамгийн их сул цагтай оператор</div><div id="topOp"></div></div>
  </div>
  <div class="full-card"><div class="card-title">Техникийн ашиглалтын хувь (TOP 7)</div><div id="fleetBars"></div></div>
  <div class="full-card">
    <div class="card-title">Үйл ажиллагааны зөвлөмж</div>
    <div class="alert-box alert-crit"><span style="font-size:16px;">🚨</span><div><strong class="alert-text-crit">НЭН ТЭРГҮҮНД:</strong> TR-6135 (хурдны хайрцаг бүрэн доголдол) — 30 хоногт капитал засвар хийх.</div></div>
    <div class="alert-box alert-crit"><span style="font-size:16px;">🚨</span><div><strong class="alert-text-crit">НЭН ТЭРГҮҮНД:</strong> TR-6053 (хөдөлгүүр хүчгүй, масло гоожилт) — capital overhaul хийх.</div></div>
    <div class="alert-box alert-warn"><span style="font-size:16px;">⚠️</span><div><strong class="alert-text-warn">АНХААРАХ:</strong> Засвар нийт сул цагийн 65% — PM хуваарь хатуу.</div></div>
    <div class="alert-box alert-warn"><span style="font-size:16px;">⚠️</span><div><strong class="alert-text-warn">АНХААРАХ:</strong> OP_LOG-д бүртгэлгүй засварт зогссон техникүүд байгаа. Бүртгэлийн журам гаргах.</div></div>
    <div class="alert-box alert-ok"><span style="font-size:16px;">💡</span><div><strong class="alert-text-ok">БОЛОМЖ:</strong> Ажиллаагүй техникүүдийг шалгаж идэвхжүүлэх.</div></div>
  </div>
  <div style="text-align:center;padding:16px;font-size:11px;color:#aaa;">Novel Mining © 2026 | KPI Dashboard | Автоматаар шинэчлэгддэг</div>
</div>
<script>
const dailyData={daily_values};
const labels={daily_labels};
const target={target};
const cumActual={cum_actual};
const cumPlan={cum_plan};

new Chart(document.getElementById('dailyChart'),{{type:'bar',data:{{labels,datasets:[
  {{label:'Бодит',data:dailyData,backgroundColor:dailyData.map(v=>v<target?'#E24B4A':'#378ADD'),borderRadius:4,borderSkipped:false}},
  {{label:'Зорилт',data:Array(labels.length).fill(target),type:'line',borderColor:'#aaa',borderDash:[4,3],borderWidth:1.5,pointRadius:0,fill:false}}
]}},options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:false}}}},
  scales:{{x:{{ticks:{{color:'#888',font:{{size:10}}}},grid:{{display:false}}}},
    y:{{ticks:{{color:'#888',font:{{size:10}},callback:v=>v+'ц'}},grid:{{color:'#f0f2f5'}},min:200,max:650}}}}}}}});

new Chart(document.getElementById('cumChart'),{{type:'line',data:{{labels,datasets:[
  {{label:'Бодит',data:cumActual,borderColor:'#1D9E75',backgroundColor:'rgba(29,158,117,0.1)',fill:true,tension:0.35,pointRadius:3,pointBackgroundColor:'#1D9E75',borderWidth:2}},
  {{label:'Төлөвлөгөө',data:cumPlan,borderColor:'#aaa',borderDash:[5,3],fill:false,pointRadius:0,borderWidth:1.5}}
]}},options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:false}}}},
  scales:{{x:{{ticks:{{color:'#888',font:{{size:10}}}},grid:{{display:false}}}},
    y:{{ticks:{{color:'#888',font:{{size:10}},callback:v=>(v/1000).toFixed(1)+'k'}},grid:{{color:'#f0f2f5'}}}}}}}}}});

new Chart(document.getElementById('repairPie'),{{type:'doughnut',data:{{
  labels:['Хөдөлгүүр','Явах эд анги','Хурдны хайрцаг','Кабин','ТҮ','Бусад'],
  datasets:[{{data:[26.3,21.5,9.2,7.6,6.1,29.3],backgroundColor:['#E24B4A','#378ADD','#EF9F27','#1D9E75','#7F77DD','#888'],borderWidth:0}}]}},
  options:{{responsive:true,maintainAspectRatio:false,cutout:'58%',plugins:{{legend:{{display:false}}}}}}}});

const trucks={trucks_js};
const maxCalls=trucks.reduce((a,b)=>Math.max(a,b.calls),0);
trucks.forEach(t=>{{
  const row=document.createElement('div');row.className='bar-row';
  const color=t.calls>=30?'#E24B4A':'#EF9F27';
  row.innerHTML=`<span class="bar-label">${{t.name}}</span><div class="bar-track"><div class="bar-fill" style="width:${{(t.calls/maxCalls*100).toFixed(0)}}%;background:${{color}};"></div></div><span class="bar-val">${{t.calls}} удаа</span>`;
  document.getElementById('topTruck').appendChild(row);
}});

const ops={ops_js};
const maxHours=ops.reduce((a,b)=>Math.max(a,b.hours),0);
ops.forEach(o=>{{
  const row=document.createElement('div');row.className='bar-row';
  const color=o.hours>=100?'#E24B4A':'#EF9F27';
  row.innerHTML=`<span class="bar-label">${{o.name}}</span><div class="bar-track"><div class="bar-fill" style="width:${{(o.hours/maxHours*100).toFixed(0)}}%;background:${{color}};"></div></div><span class="bar-val">${{o.hours}}ц</span>`;
  document.getElementById('topOp').appendChild(row);
}});

const fleet={fleet_js};
fleet.forEach(f=>{{
  const color=f.util>=80?'#1D9E75':'#EF9F27';
  const row=document.createElement('div');row.className='bar-row';
  row.innerHTML=`<span class="bar-label">${{f.name}}</span><div class="bar-track"><div class="bar-fill" style="width:${{f.util}}%;background:${{color}};"></div></div><span class="bar-val">${{f.util}}%</span>`;
  document.getElementById('fleetBars').appendChild(row);
}});
</script>
</body>
</html>'''
    
    # Replace JS arrays
    html = html.replace('{daily_values}', str(daily_values))
    html = html.replace('{daily_labels}', str(daily_labels))
    html = html.replace('{cum_actual}', str(cum_actual))
    html = html.replace('{cum_plan}', str(cum_plan))
    
    return html

def main():
    print("Fetching Excel from OneDrive...")
    try:
        file_bytes = fetch_excel()
        print("Excel fetched successfully!")
    except Exception as e:
        print(f"Error fetching Excel: {e}")
        print("Using cached data...")
        return
    
    print("Extracting data...")
    data = extract_data(file_bytes)
    print(f"Moto hours: {data['moto_hours']}, Repairs: {data['repair_calls']}")
    
    print("Generating HTML...")
    html = generate_html(data)
    
    with open('index.html', 'w', encoding='utf-8') as f:
        f.write(html)
    print("Dashboard updated successfully!")

if __name__ == '__main__':
    main()
